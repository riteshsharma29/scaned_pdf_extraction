import fitz
from tesserocr import PyTessBaseAPI
import PyPDF2
import os
from natsort import natsorted, ns
import xlsxwriter
from pandas import ExcelWriter
from openpyxl import load_workbook
import codecs
import pandas as pd


#
# # print(tesserocr.tesseract_version())  # print tesseract-ocr version
# # print(tesserocr.get_languages())  # prints tessdata path and list of available languages
#
def extract_pdf_img(pdf_file):
    pdfReader = PyPDF2.PdfFileReader(pdf_file)
    # count number of pages
    totalPages = pdfReader.numPages
    for page in range(0, totalPages):
        pdffile = "pdf_scanned_ocr.pdf"
        doc = fitz.open(pdffile)
        page = doc.load_page(page)  # number of page
        pix = page.get_pixmap()
        output = "{}.png".format(page)
        output = output.replace(" ", "")
        pix.save(os.path.join("pdfimg", output))
        doc.close()

# call extract img from pdf function
extract_pdf_img("pdf_scanned_ocr.pdf")

# sort image list alphanumerically
images = os.listdir("pdfimg")
fileList = natsorted(images, key=lambda y: y.lower())#sort alphanumeric in ascending order

# generate empty text files
def write_text_file(txt_file,extracted_text):
    f = codecs.open(txt_file, "w", encoding="utf-8")
    f.write(extracted_text)

# read pdf img
def read_pdf_img():
    with PyTessBaseAPI(lang='eng', path='testdata\\') as api:
        for img in fileList:
            try:
                api.SetImageFile(os.path.join("pdfimg",img))
            # print(api.GetUTF8Text())
                write_text_file(os.path.join("pdfimgtxt",img)+ ".txt",api.GetUTF8Text())
            except:
                continue

# call pdf img function
read_pdf_img()


# save extracted pdf in excel

# Create Excel Workbook
LOG_File = os.path.join('output.xlsx')
workbook = xlsxwriter.Workbook(LOG_File)
worksheet = workbook.add_worksheet()
workbook.close()

# Load excel Workbook using openpyxl
book = load_workbook(LOG_File)
writer = ExcelWriter(LOG_File, engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

text_files = os.listdir("pdfimgtxt")
text_files_list = natsorted(text_files, key=lambda y: y.lower())#sort alphanumeric in ascending order

for  txt_file in text_files_list:
    datafile = codecs.open(os.path.join("pdfimgtxt",txt_file), "r", encoding='utf-8')
    df = pd.DataFrame(columns=['Lines'])

    for lines in datafile:
        df.loc[df.shape[0]] = lines.strip('\n\r')
    df.to_excel(writer, sheet_name=txt_file[0:6], index=False)

if len(book.sheetnames) > 1:
    first_sheet = book['Sheet1']
    book.remove(first_sheet)

writer.save()
